#!/usr/bin/env python3
"""
Debug script to investigate date formatting issues in Excel generation
"""

import pandas as pd
from datetime import datetime, date
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path

# Crear un workbook de prueba
wb = openpyxl.Workbook()
wb.create_sheet("Test Dates")
ws = wb["Test Dates"]

# Probar diferentes tipos de dates
test_dates = [
    ("String date", "2025-05-15"),
    ("Python datetime", datetime(2025, 5, 15)),
    ("Python date", date(2025, 5, 15)),
    ("Pandas timestamp", pd.Timestamp('2025-05-15')),
]

# Escribir diferentes tipos de fechas
row = 1
for desc, fecha_value in test_dates:
    ws[f'A{row}'] = desc
    
    # Celda para la fecha
    fecha_cell = ws[f'B{row}']
    
    try:
        # Aplicar la misma lógica que el código original
        if hasattr(fecha_value, 'to_pydatetime'):
            fecha_cell.value = fecha_value.to_pydatetime()
            print(f"{desc}: Convertido con to_pydatetime() -> {fecha_cell.value} (tipo: {type(fecha_cell.value)})")
        elif hasattr(fecha_value, 'date'):
            fecha_cell.value = fecha_value
            print(f"{desc}: Usado directamente (tiene .date) -> {fecha_cell.value} (tipo: {type(fecha_cell.value)})")
        elif isinstance(fecha_value, str):
            fecha_dt = datetime.strptime(fecha_value, '%Y-%m-%d')
            fecha_cell.value = fecha_dt
            print(f"{desc}: Convertido desde string -> {fecha_cell.value} (tipo: {type(fecha_cell.value)})")
        elif isinstance(fecha_value, date):
            fecha_cell.value = datetime.combine(fecha_value, datetime.min.time())
            print(f"{desc}: Convertido desde date -> {fecha_cell.value} (tipo: {type(fecha_cell.value)})")
        else:
            fecha_cell.value = fecha_value
            print(f"{desc}: Usado directamente -> {fecha_cell.value} (tipo: {type(fecha_cell.value)})")
        
        # Aplicar formato de fecha
        fecha_cell.number_format = 'dd/mm/yyyy'
        print(f"  → Formato aplicado: {fecha_cell.number_format}")
        
    except Exception as e:
        print(f"❌ Error con {desc}: {e}")
        fecha_cell.value = str(fecha_value)
    
    row += 1

# Guardar archivo de prueba
test_file = "/Users/alecarrasco/Documents/06_DESARROLLOS/pago_publicidad/asrun-report/test_date_debug.xlsx"
wb.save(test_file)
print(f"\n📁 Archivo de prueba guardado: {test_file}")

# Verificar el archivo generado
print("\n🔍 Verificando el archivo generado...")
wb_check = openpyxl.load_workbook(test_file)
ws_check = wb_check["Test Dates"]  # Usar el nombre específico de la hoja

for row_num in range(1, 5):
    cell = ws_check[f'B{row_num}']
    print(f"Fila {row_num}: Valor={cell.value}, Tipo={type(cell.value)}, Formato={cell.number_format}")
