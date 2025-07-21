#!/usr/bin/env python3
"""
Consultor de base de datos AsRun
Permite consultar y analizar datos almacenados en la base de datos SQLite
"""

import pandas as pd
from datetime import datetime, date, timedelta
from pathlib import Path
import json
import sys
import os

# Add the current directory to the path for absolute imports
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from database_manager import AsRunDatabase


class AsRunConsultor:
    """Clase para consultar y analizar datos de la base de datos AsRun"""
    
    def __init__(self, db_path=None):
        """Inicializar el consultor con la base de datos"""
        if db_path is None:
            # Usar SIEMPRE la base de datos principal en el directorio raíz del proyecto
            current_dir = Path(__file__).parent
            # Prioridad única: BD principal en directorio padre (raíz del proyecto)
            db_path = str(current_dir.parent / "asrun_database.db")
        
        self.db = AsRunDatabase(db_path)
        self.directorio_proyecto = Path(db_path).parent
        print(f"🗄️  Conectado a la base de datos: {db_path}")
    
    def mostrar_estadisticas_generales(self):
        """Mostrar estadísticas generales de la base de datos"""
        print("\n📊 ESTADÍSTICAS GENERALES")
        print("=" * 50)
        
        stats = self.db.obtener_estadisticas_generales()
        print(f"📺 Total de emisiones registradas: {stats['total_emisiones']:,}")
        print(f"👥 Total de clientes únicos: {stats['total_clientes']}")
        print(f"📄 Total de reportes generados: {stats['total_reportes']}")
        
        if stats['fecha_inicio'] and stats['fecha_fin']:
            print(f"📅 Rango de fechas: {stats['fecha_inicio']} → {stats['fecha_fin']}")
            
            fecha_inicio = datetime.strptime(stats['fecha_inicio'], '%Y-%m-%d').date()
            fecha_fin = datetime.strptime(stats['fecha_fin'], '%Y-%m-%d').date()
            dias_total = (fecha_fin - fecha_inicio).days + 1
            print(f"📆 Total de días cubiertos: {dias_total}")
        
        return stats
    
    def mostrar_resumen_por_cliente(self, fecha_inicio=None, fecha_fin=None, top_n=None):
        """Mostrar resumen de emisiones agrupadas por cliente"""
        print("\n👥 RESUMEN POR CLIENTE")
        print("=" * 50)
        
        if fecha_inicio or fecha_fin:
            periodo = f" (Período: {fecha_inicio or 'inicio'} → {fecha_fin or 'fin'})"
            print(f"Filtro aplicado{periodo}")
            print()
        
        df_resumen = self.db.obtener_resumen_por_cliente(fecha_inicio, fecha_fin)
        
        if df_resumen.empty:
            print("❌ No se encontraron datos para el período especificado")
            return
        
        # Limitar a top N si se especifica
        if top_n and top_n > 0:
            df_resumen = df_resumen.head(top_n)
            print(f"📋 Mostrando top {top_n} clientes:")
        else:
            print(f"📋 Todos los clientes ({len(df_resumen)} total):")
        
        print()
        for _, fila in df_resumen.iterrows():
            print(f"🎯 {fila['cliente']}")
            print(f"   • Emisiones: {fila['total_emisiones']:,}")
            print(f"   • Días activos: {fila['dias_activos']}")
            print(f"   • Primera emisión: {fila['primera_emision']}")
            print(f"   • Última emisión: {fila['ultima_emision']}")
            print()
        
        return df_resumen
    
    def consultar_cliente_especifico(self, cliente, fecha_inicio=None, fecha_fin=None, mostrar_detalle=False):
        """Consultar emisiones de un cliente específico"""
        print(f"\n🎯 CONSULTA ESPECÍFICA - Cliente: {cliente}")
        print("=" * 50)
        
        if fecha_inicio or fecha_fin:
            periodo = f" (Período: {fecha_inicio or 'inicio'} → {fecha_fin or 'fin'})"
            print(f"Filtro aplicado{periodo}")
            print()
        
        emisiones = self.db.consultar_emisiones_por_cliente(cliente, fecha_inicio, fecha_fin)
        
        if emisiones.empty:
            print(f"❌ No se encontraron emisiones para el cliente '{cliente}' en el período especificado")
            return
        
        print(f"📊 Resumen del cliente '{cliente}':")
        print(f"   • Total de emisiones: {len(emisiones):,}")
        print(f"   • Rango de fechas: {emisiones['dia_emision'].min()} → {emisiones['dia_emision'].max()}")
        print(f"   • Horario típico: {emisiones['hora_emision'].min()} → {emisiones['hora_emision'].max()}")
        
        # Análisis de status
        if 'status' in emisiones.columns:
            status_counts = emisiones['status'].value_counts()
            print(f"   • Status de emisiones:")
            for status, count in status_counts.items():
                print(f"     - {status}: {count} emisiones ({count/len(emisiones)*100:.1f}%)")
        
        # Mostrar detalle si se solicita
        if mostrar_detalle:
            print(f"\n📋 Detalle de emisiones (últimas 10):")
            print("-" * 80)
            
            # Mostrar últimas 10 emisiones
            ultimas_emisiones = emisiones.sort_values(['dia_emision', 'hora_emision']).tail(10)
            
            for _, emision in ultimas_emisiones.iterrows():
                status_str = f" [{emision.get('status', 'N/A')}]" if 'status' in emisiones.columns else ""
                print(f"📅 {emision['dia_emision']} {emision['hora_emision']} - {emision['titulo']}{status_str}")
        
        return emisiones
    
    def analizar_periodo(self, fecha_inicio, fecha_fin):
        """Analizar emisiones en un período específico"""
        print(f"\n📅 ANÁLISIS DE PERÍODO: {fecha_inicio} → {fecha_fin}")
        print("=" * 60)
        
        emisiones = self.db.obtener_emisiones_por_fecha(fecha_inicio, fecha_fin)
        
        if emisiones.empty:
            print(f"❌ No se encontraron emisiones en el período {fecha_inicio} → {fecha_fin}")
            return
        
        print(f"📊 Resumen del período:")
        print(f"   • Total de emisiones: {len(emisiones):,}")
        print(f"   • Clientes únicos: {emisiones['cliente'].nunique()}")
        print(f"   • Días con actividad: {emisiones['dia_emision'].nunique()}")
        
        # Análisis por cliente
        resumen_clientes = emisiones.groupby('cliente').agg({
            'titulo': 'count',
            'dia_emision': ['min', 'max', 'nunique']
        }).round(2)
        
        resumen_clientes.columns = ['Emisiones', 'Primer_Día', 'Último_Día', 'Días_Activos']
        resumen_clientes = resumen_clientes.sort_values('Emisiones', ascending=False)
        
        print(f"\n👥 Top 10 clientes en el período:")
        for cliente, datos in resumen_clientes.head(10).iterrows():
            print(f"🎯 {cliente}")
            print(f"   • Emisiones: {datos['Emisiones']:,}")
            print(f"   • Días activos: {datos['Días_Activos']}")
            print(f"   • Período: {datos['Primer_Día']} → {datos['Último_Día']}")
            print()
        
        # Análisis de status si está disponible
        if 'status' in emisiones.columns:
            print(f"\n📊 Análisis de STATUS:")
            status_counts = emisiones['status'].value_counts()
            for status, count in status_counts.items():
                print(f"   • {status}: {count} emisiones ({count/len(emisiones)*100:.1f}%)")
        
        return emisiones
    
    def buscar_por_titulo(self, titulo, fecha_inicio=None, fecha_fin=None):
        """Buscar emisiones por título o programa"""
        print(f"\n🔍 BÚSQUEDA POR TÍTULO: '{titulo}'")
        print("=" * 50)
        
        # Implementar búsqueda usando SQL LIKE
        emisiones = self._buscar_por_titulo_sql(titulo, fecha_inicio, fecha_fin)
        
        if emisiones.empty:
            print(f"❌ No se encontraron emisiones con el título '{titulo}'")
            return
        
        print(f"📊 Resultados de búsqueda:")
        print(f"   • Emisiones encontradas: {len(emisiones):,}")
        print(f"   • Clientes únicos: {emisiones['cliente'].nunique()}")
        print(f"   • Rango de fechas: {emisiones['dia_emision'].min()} → {emisiones['dia_emision'].max()}")
        
        # Mostrar resumen por cliente
        resumen_clientes = emisiones.groupby('cliente').size().sort_values(ascending=False)
        
        print(f"\n👥 Distribución por cliente:")
        for cliente, cantidad in resumen_clientes.items():
            print(f"   • {cliente}: {cantidad} emisiones")
        
        # Mostrar últimas 5 emisiones como ejemplo
        print(f"\n📋 Últimas 5 emisiones encontradas:")
        print("-" * 80)
        
        ultimas_emisiones = emisiones.sort_values(['dia_emision', 'hora_emision']).tail(5)
        
        for _, emision in ultimas_emisiones.iterrows():
            status_str = f" [{emision.get('status', 'N/A')}]" if 'status' in emisiones.columns else ""
            print(f"📅 {emision['dia_emision']} {emision['hora_emision']} - {emision['cliente']} - {emision['titulo']}{status_str}")
        
        return emisiones
    
    def _buscar_por_titulo_sql(self, titulo, fecha_inicio=None, fecha_fin=None):
        """Buscar emisiones por título usando SQL LIKE"""
        try:
            # Construir la consulta SQL
            query = """
                SELECT * FROM emisiones 
                WHERE titulo LIKE ? COLLATE NOCASE
            """
            params = [f"%{titulo}%"]
            
            # Agregar filtros de fecha si se proporcionan
            if fecha_inicio:
                query += " AND dia_emision >= ?"
                params.append(fecha_inicio)
            
            if fecha_fin:
                query += " AND dia_emision <= ?"
                params.append(fecha_fin)
            
            query += " ORDER BY dia_emision DESC, hora_emision DESC"
            
            # Ejecutar consulta usando pandas para obtener DataFrame
            import sqlite3
            with sqlite3.connect(self.db.db_path) as conn:
                df_resultado = pd.read_sql_query(query, conn, params=tuple(params))
            
            return df_resultado
            
        except Exception as e:
            print(f"Error en búsqueda por título: {e}")
            return pd.DataFrame()
    
    def generar_reporte_desde_consulta(self, fecha_inicio=None, fecha_fin=None, cliente=None):
        """Generar reporte personalizado basado en filtros de consulta"""
        try:
            print(f"\n📋 GENERANDO REPORTE PERSONALIZADO")
            print("=" * 50)
            
            # Configurar filtros
            filtros = []
            if fecha_inicio:
                filtros.append(f"Desde: {fecha_inicio}")
            if fecha_fin:
                filtros.append(f"Hasta: {fecha_fin}")
            if cliente:
                filtros.append(f"Cliente: {cliente}")
            
            filtros_str = " | ".join(filtros) if filtros else "Sin filtros"
            print(f"🔍 Filtros aplicados: {filtros_str}")
            
            # Obtener emisiones con filtros
            if cliente and cliente != "Todos":
                emisiones = self.db.consultar_emisiones_por_cliente(cliente, fecha_inicio, fecha_fin)
            else:
                emisiones = self.db.obtener_emisiones_por_fecha(fecha_inicio, fecha_fin)
            
            if emisiones.empty:
                print("❌ No se encontraron datos para generar el reporte")
                return None
            
            print(f"📊 Datos encontrados: {len(emisiones):,} emisiones")
            print(f"👥 Clientes únicos: {emisiones['cliente'].nunique()}")
            
            # Definir directorio de reportes
            directorio_reportes = self.directorio_proyecto / "reportes"
            directorio_reportes.mkdir(exist_ok=True)
            
            # Generar nombre base del archivo
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            
            if cliente and cliente != "Todos":
                nombre_base = f"reporte_{cliente.replace(' ', '_')}_{timestamp}"
            else:
                nombre_base = f"reporte_personalizado_{timestamp}"
            
            archivos_generados = []
            
            # Generar archivo TXT
            nombre_txt, ruta_txt = self._obtener_nombre_con_version(directorio_reportes, nombre_base, ".txt")
            if self._generar_archivo_txt(emisiones, ruta_txt, filtros_str):
                archivos_generados.append(ruta_txt)
                print(f"✅ Archivo TXT generado: {nombre_txt}")
            
            # Generar archivo Excel
            nombre_excel, ruta_excel = self._obtener_nombre_con_version(directorio_reportes, nombre_base, ".xlsx")
            if self._generar_archivo_excel(emisiones, ruta_excel):
                archivos_generados.append(ruta_excel)
                print(f"✅ Archivo Excel generado: {nombre_excel}")
            
            print(f"\n🎯 Reporte generado exitosamente:")
            print(f"📁 Directorio: {directorio_reportes}")
            print(f"📄 Archivos: {len(archivos_generados)}")
            
            return {
                'success': True,
                'archivos': archivos_generados,
                'total_emisiones': len(emisiones),
                'clientes_unicos': emisiones['cliente'].nunique()
            }
            
        except Exception as e:
            print(f"❌ Error al generar reporte: {e}")
            return None
    
    def _generar_archivo_txt(self, emisiones, ruta_archivo, filtros_str=""):
        """Generar archivo de reporte en formato TXT"""
        try:
            with open(ruta_archivo, 'w', encoding='utf-8') as f:
                # Encabezado del reporte
                f.write("=" * 80 + "\n")
                f.write("REPORTE DE EMISIONES ASRUN\n")
                f.write("=" * 80 + "\n")
                f.write(f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write(f"Filtros aplicados: {filtros_str}\n")
                f.write(f"Total de emisiones: {len(emisiones):,}\n")
                f.write(f"Clientes únicos: {emisiones['cliente'].nunique()}\n")
                f.write("=" * 80 + "\n\n")
                
                # Resumen por cliente
                resumen_clientes = emisiones.groupby('cliente').agg({
                    'titulo': 'count',
                    'dia_emision': ['min', 'max']
                }).round(2)
                
                resumen_clientes.columns = ['Total_Emisiones', 'Primer_Día', 'Último_Día']
                resumen_clientes = resumen_clientes.sort_values('Total_Emisiones', ascending=False)
                
                f.write("RESUMEN POR CLIENTE\n")
                f.write("-" * 50 + "\n")
                
                for cliente, datos in resumen_clientes.iterrows():
                    f.write(f"Cliente: {cliente}\n")
                    f.write(f"  Emisiones: {datos['Total_Emisiones']:,}\n")
                    f.write(f"  Período: {datos['Primer_Día']} → {datos['Último_Día']}\n\n")
                
                # Análisis de STATUS si está disponible
                if 'status' in emisiones.columns:
                    f.write("\nANÁLISIS DE STATUS\n")
                    f.write("-" * 30 + "\n")
                    
                    status_counts = emisiones['status'].value_counts()
                    for status, count in status_counts.items():
                        porcentaje = (count / len(emisiones)) * 100
                        f.write(f"{status}: {count} emisiones ({porcentaje:.1f}%)\n")
                    
                    # Detalles de Lost XPoint Path si existe
                    lost_xpoint_df = emisiones[emisiones['status'].str.contains('Lost XPoint Path', case=False, na=False)]
                    if not lost_xpoint_df.empty:
                        f.write(f"\nDETALLE DE LOST XPOINT PATH ({len(lost_xpoint_df)} registros)\n")
                        f.write("-" * 50 + "\n")
                        
                        for _, row in lost_xpoint_df.iterrows():
                            # Formatear hora sin decimales
                            hora_limpia = self._formatear_hora_sin_decimales(row['hora_emision'])
                            f.write(f"Fecha: {row['dia_emision']} | Hora: {hora_limpia} | Cliente: {row['cliente']}\n")
                            f.write(f"Título: {row['titulo']}\n")
                            if 'media_id' in row and pd.notna(row['media_id']):
                                f.write(f"Media ID: {row['media_id']}\n")
                            f.write("\n")
                
                f.write("\n" + "=" * 80 + "\n")
                f.write("FIN DEL REPORTE\n")
                f.write("=" * 80 + "\n")
            
            return True
            
        except Exception as e:
            print(f"❌ Error generando archivo TXT: {e}")
            return False
    
    def _generar_archivo_excel(self, emisiones, ruta_archivo_xlsx):
        """Generar archivo de reporte en formato Excel con múltiples hojas"""
        try:
            # Preparar datos para Excel
            df_excel = emisiones.copy()
            
            # Convertir fechas para formato Excel
            df_excel['Fecha'] = pd.to_datetime(df_excel['dia_emision']).dt.strftime('%Y-%m-%d')
            # Aplicar formateo de hora sin decimales
            df_excel['Hora'] = df_excel['hora_emision'].apply(self._formatear_hora_sin_decimales)
            
            # Seleccionar y reordenar columnas
            columnas_base = ['Fecha', 'Hora', 'cliente', 'titulo', 'duracion']
            if 'media_id' in df_excel.columns:
                columnas_base.append('media_id')
            if 'status' in df_excel.columns:
                columnas_base.append('status')
            
            columnas_existentes = [col for col in columnas_base if col in df_excel.columns]
            df_excel = df_excel[columnas_existentes]
            
            # Renombrar columnas para mejor presentación
            nombres_columnas = {
                'cliente': 'Cliente',
                'titulo': 'Título',
                'duracion': 'Duración',
                'media_id': 'Media ID',
                'status': 'Status'
            }
            df_excel = df_excel.rename(columns=nombres_columnas)
            
            with pd.ExcelWriter(ruta_archivo_xlsx, engine='openpyxl') as writer:
                # Hoja 1: Datos principales
                df_excel.to_excel(writer, sheet_name='Emisiones', index=False)
                
                # Hoja 2: Resumen por cliente
                resumen_cliente = df_excel.groupby('Cliente').agg({
                    'Título': 'count',
                    'Fecha': ['min', 'max']
                }).round(2)
                resumen_cliente.columns = ['Total Emisiones', 'Primera Fecha', 'Última Fecha']
                resumen_cliente.to_excel(writer, sheet_name='Resumen por Cliente')
                
                # Hoja 3: Resumen por fecha
                resumen_fecha = df_excel.groupby('Fecha').agg({
                    'Título': 'count',
                    'Cliente': 'nunique'
                }).rename(columns={'Título': 'Total Emisiones', 'Cliente': 'Clientes Únicos'})
                resumen_fecha.to_excel(writer, sheet_name='Resumen por Fecha')
                
                # Hoja 4: Lost XPoint Path Analysis (si hay datos de status)
                if 'Status' in df_excel.columns:
                    lost_xpoint_df = self._analizar_lost_xpoint_path(emisiones)
                    if not lost_xpoint_df.empty:
                        lost_xpoint_df.to_excel(writer, sheet_name='Lost XPoint Path', index=False)
                        print(f"   📊 Hoja 'Lost XPoint Path' agregada: {len(lost_xpoint_df)} registros")
                        
                        # Formatear la hoja Lost XPoint Path
                        try:
                            worksheet_lost = writer.sheets['Lost XPoint Path']
                            
                            # Configurar ancho de columnas específicos para Lost XPoint Path usando openpyxl
                            worksheet_lost.column_dimensions['A'].width = 12  # Fecha
                            worksheet_lost.column_dimensions['B'].width = 20  # Cliente
                            worksheet_lost.column_dimensions['C'].width = 12  # Hora Inicio
                            worksheet_lost.column_dimensions['D'].width = 30  # Título/Programa
                            worksheet_lost.column_dimensions['E'].width = 25  # Media ID
                            worksheet_lost.column_dimensions['F'].width = 15  # Duración
                        except Exception as format_error:
                            print(f"   ⚠️  No se pudo aplicar formato a hoja Lost XPoint Path: {format_error}")
                
                # Configurar formato de las hojas principales
                try:
                    # Para openpyxl, aplicar formato básico
                    from openpyxl.styles import Font, PatternFill, Border, Side
                    
                    # Aplicar formato a la hoja principal
                    worksheet = writer.sheets['Emisiones']
                    
                    # Formato para encabezados
                    header_font = Font(bold=True)
                    header_fill = PatternFill(start_color='D7E4BC', end_color='D7E4BC', fill_type='solid')
                    thin_border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                    
                    # Aplicar formato a encabezados
                    for col_num, value in enumerate(df_excel.columns.values, 1):
                        cell = worksheet.cell(row=1, column=col_num)
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.border = thin_border
                    
                    # Ajustar ancho de columnas principales
                    worksheet.column_dimensions['A'].width = 12  # Fecha
                    worksheet.column_dimensions['B'].width = 10  # Hora
                    worksheet.column_dimensions['C'].width = 20  # Cliente
                    worksheet.column_dimensions['D'].width = 40  # Título
                    worksheet.column_dimensions['E'].width = 12  # Duración
                    if 'Media ID' in df_excel.columns:
                        worksheet.column_dimensions['F'].width = 25  # Media ID
                    if 'Status' in df_excel.columns:
                        col_idx = list(df_excel.columns).index('Status')
                        col_letter = chr(65 + col_idx)  # Convertir a letra (A, B, C...)
                        worksheet.column_dimensions[col_letter].width = 20  # Status
                        
                except Exception as format_error:
                    print(f"   ⚠️  No se pudo aplicar formato avanzado: {format_error}")
                    # El archivo Excel se creará sin formato especial
            
            return True
            
        except Exception as e:
            print(f"❌ Error generando archivo Excel: {e}")
            return False
    
    def _analizar_lost_xpoint_path(self, emisiones):
        """Analizar y extraer registros con Lost XPoint Path"""
        try:
            # Filtrar registros con Lost XPoint Path
            if 'status' not in emisiones.columns:
                print("   ⚠️  Columna 'status' no encontrada en los datos")
                return pd.DataFrame()
            
            lost_xpoint_mask = emisiones['status'].str.contains('Lost XPoint Path', case=False, na=False)
            lost_xpoint_df = emisiones[lost_xpoint_mask].copy()
            
            if len(lost_xpoint_df) == 0:
                print("   ℹ️  No se encontraron registros con 'Lost XPoint Path'")
                return pd.DataFrame()
            
            # Preparar datos para Excel con formato específico
            df_excel = lost_xpoint_df.copy()
            
            # Convertir fechas y horas
            df_excel['Fecha'] = pd.to_datetime(df_excel['dia_emision']).dt.strftime('%Y-%m-%d')
            
            # Aplicar formateo de hora sin decimales para Lost XPoint Path
            df_excel['Hora_Inicio'] = df_excel['hora_emision'].apply(self._formatear_hora_sin_decimales)
            
            # Seleccionar y reordenar columnas para el reporte (sin Hora Fin ni Duración Calculada)
            columnas_reporte = ['Fecha', 'cliente', 'Hora_Inicio', 'titulo', 'media_id', 'duracion']
            columnas_existentes = [col for col in columnas_reporte if col in df_excel.columns]
            
            df_resultado = df_excel[columnas_existentes].copy()
            
            # Renombrar columnas para el reporte
            nombres_finales = ['Fecha', 'Cliente', 'Hora Inicio', 'Título/Programa', 'Media ID', 'Duración']
            df_resultado.columns = nombres_finales[:len(df_resultado.columns)]
            
            # Ordenar por fecha y hora
            df_resultado = df_resultado.sort_values(['Fecha', 'Hora Inicio'], ascending=[True, True])
            df_resultado = df_resultado.reset_index(drop=True)
            
            print(f"   🔍 Análisis Lost XPoint Path: {len(df_resultado)} registros encontrados")
            return df_resultado
            
        except Exception as e:
            print(f"   ❌ Error analizando Lost XPoint Path: {e}")
            return pd.DataFrame()
    
    def _calcular_hora_fin(self, hora_inicio_series, duracion_series):
        """Calcular hora de fin basada en hora de inicio y duración"""
        horas_fin = []
        
        for hora_inicio, duracion in zip(hora_inicio_series, duracion_series):
            try:
                # Parsear hora de inicio
                if pd.isna(hora_inicio):
                    horas_fin.append("N/A")
                    continue
                
                hora_inicio_str = str(hora_inicio).strip()
                if not hora_inicio_str:
                    horas_fin.append("N/A")
                    continue
                
                # Parsear duración (formato típico: 00:01:30 o similar)
                duracion_str = str(duracion).strip() if not pd.isna(duracion) else "00:00:00"
                
                # Convertir hora de inicio
                try:
                    inicio_dt = datetime.strptime(hora_inicio_str, '%H:%M:%S')
                except ValueError:
                    try:
                        inicio_dt = datetime.strptime(hora_inicio_str, '%H:%M')
                    except ValueError:
                        horas_fin.append("Formato inválido")
                        continue
                
                # Convertir duración
                try:
                    # Intentar formato HH:MM:SS
                    duracion_parts = duracion_str.split(':')
                    if len(duracion_parts) >= 3:
                        horas = int(duracion_parts[0])
                        minutos = int(duracion_parts[1])
                        segundos = int(duracion_parts[2])
                        duracion_delta = timedelta(hours=horas, minutes=minutos, seconds=segundos)
                    elif len(duracion_parts) == 2:
                        # Formato MM:SS
                        minutos = int(duracion_parts[0])
                        segundos = int(duracion_parts[1])
                        duracion_delta = timedelta(minutes=minutos, seconds=segundos)
                    else:
                        # Asumir que es solo minutos
                        minutos = int(float(duracion_str))
                        duracion_delta = timedelta(minutes=minutos)
                        
                except (ValueError, IndexError):
                    horas_fin.append("Duración inválida")
                    continue
                
                # Calcular hora de fin
                fin_dt = inicio_dt + duracion_delta
                horas_fin.append(fin_dt.strftime('%H:%M:%S'))
                
            except Exception as e:
                horas_fin.append(f"Error: {str(e)[:10]}")
        
        return horas_fin
    
    def _calcular_duracion_evento(self, hora_inicio_series, hora_fin_series):
        """Calcular duración de eventos Lost XPoint Path"""
        duraciones = []
        
        for hora_inicio, hora_fin in zip(hora_inicio_series, hora_fin_series):
            try:
                if pd.isna(hora_inicio) or pd.isna(hora_fin) or hora_inicio == "N/A" or hora_fin == "N/A":
                    duraciones.append("N/A")
                    continue
                
                inicio_dt = datetime.strptime(str(hora_inicio), '%H:%M:%S')
                fin_dt = datetime.strptime(str(hora_fin), '%H:%M:%S')
                
                # Manejar caso donde la hora de fin es al día siguiente
                if fin_dt < inicio_dt:
                    fin_dt += timedelta(days=1)
                
                duracion = fin_dt - inicio_dt
                
                # Formatear duración como HH:MM:SS
                total_seconds = int(duracion.total_seconds())
                hours = total_seconds // 3600
                minutes = (total_seconds % 3600) // 60
                seconds = total_seconds % 60
                
                duraciones.append(f"{hours:02d}:{minutes:02d}:{seconds:02d}")
                
            except (ValueError, TypeError) as e:
                duraciones.append("Error cálculo")
        
        return duraciones
    
    def _formatear_hora_sin_decimales(self, hora_str):
        """Formatear hora eliminando decimales y manteniendo solo HH:MM:SS"""
        try:
            if pd.isna(hora_str):
                return "N/A"
            
            hora_str = str(hora_str)
            
            # Si la hora tiene decimales, eliminarlos
            if '.' in hora_str:
                hora_str = hora_str.split('.')[0]
            
            # Verificar si ya está en formato HH:MM:SS
            if len(hora_str.split(':')) == 3:
                try:
                    # Validar que sea un tiempo válido
                    datetime.strptime(hora_str, '%H:%M:%S')
                    return hora_str
                except ValueError:
                    pass
            
            # Si es un objeto datetime, convertir a string sin decimales
            try:
                if isinstance(hora_str, str) and 'T' in hora_str:
                    # Formato datetime ISO
                    dt = datetime.fromisoformat(hora_str.replace('T', ' '))
                    return dt.strftime('%H:%M:%S')
                else:
                    # Intentar parsear como datetime
                    dt = pd.to_datetime(hora_str)
                    return dt.strftime('%H:%M:%S')
            except:
                return str(hora_str)[:8] if len(str(hora_str)) >= 8 else str(hora_str)
                
        except Exception as e:
            return str(hora_str)
    
    def _obtener_nombre_con_version(self, directorio, nombre_base, extension=""):
        """
        Obtener nombre de archivo con versionado automático
        
        Args:
            directorio: Path del directorio donde se guardará el archivo
            nombre_base: Nombre base del archivo (sin extensión)
            extension: Extensión del archivo (ej: '.txt', '.xlsx')
            
        Returns:
            tuple: (nombre_con_version, ruta_completa)
        """
        version = 1
        
        while True:
            nombre_versionado = f"{nombre_base}_v{version}{extension}"
            ruta_archivo = directorio / nombre_versionado
            
            # Para reportes Excel/TXT, verificar que no existan ambos formatos
            if extension in ['.txt', '.xlsx']:
                nombre_base_v = f"{nombre_base}_v{version}"
                ruta_txt = directorio / f"{nombre_base_v}.txt"
                ruta_xlsx = directorio / f"{nombre_base_v}.xlsx"
                
                if not ruta_txt.exists() and not ruta_xlsx.exists():
                    break
            else:
                # Para otros tipos de archivo, verificar solo la extensión específica
                if not ruta_archivo.exists():
                    break
            
            version += 1
        
        return nombre_versionado, ruta_archivo
    
    def preguntar_generar_reporte(self):
        """Preguntar al usuario si desea generar un reporte"""
        print("\n" + "🎯" * 30)
        respuesta = input("📝 ¿Deseas generar un reporte con estos datos? (s/N): ").strip().lower()
        
        if respuesta in ['s', 'si', 'sí', 'y', 'yes']:
            print("\n📋 Configurar filtros para el reporte:")
            
            # Pedir filtros adicionales
            fecha_inicio = input("Fecha inicio (YYYY-MM-DD) [Enter para omitir]: ").strip() or None
            fecha_fin = input("Fecha fin (YYYY-MM-DD) [Enter para omitir]: ").strip() or None
            cliente = input("Filtrar por cliente [Enter para todos]: ").strip() or None
            
            # Generar el reporte
            self.generar_reporte_desde_consulta(
                fecha_inicio=fecha_inicio,
                fecha_fin=fecha_fin,
                cliente=cliente
            )
        else:
            print("📋 No se generará reporte.")

    def menu_principal(self):
        """Menú principal del consultor"""
        print("\n" + "=" * 60)
        print("🗄️  CONSULTOR DE BASE DE DATOS ASRUN")
        print("=" * 60)
        
        while True:
            print("\n📋 Opciones disponibles:")
            print("1. 📊 Mostrar estadísticas generales")
            print("2. 👥 Resumen por cliente")
            print("3. 🎯 Consultar cliente específico")
            print("4. 📅 Analizar período específico")
            print("5. 🔍 Buscar por título/programa")
            print("6. 📝 Generar reporte personalizado")
            print("7. 🚪 Salir")
            
            try:
                opcion = input("\n🎯 Selecciona una opción (1-7): ")
                
                if opcion == '1':
                    self.mostrar_estadisticas_generales()
                elif opcion == '2':
                    fecha_inicio = input("Fecha inicio (YYYY-MM-DD) [Enter para omitir]: ").strip() or None
                    fecha_fin = input("Fecha fin (YYYY-MM-DD) [Enter para omitir]: ").strip() or None
                    top_n = input("Mostrar top N clientes (dejar en blanco para todos): ").strip()
                    top_n = int(top_n) if top_n.isdigit() else None
                    self.mostrar_resumen_por_cliente(fecha_inicio, fecha_fin, top_n)
                elif opcion == '3':
                    cliente = input("Nombre del cliente: ").strip()
                    fecha_inicio = input("Fecha inicio (YYYY-MM-DD) [Enter para omitir]: ").strip() or None
                    fecha_fin = input("Fecha fin (YYYY-MM-DD) [Enter para omitir]: ").strip() or None
                    mostrar_detalle = input("¿Mostrar detalle de emisiones? (s/N): ").strip().lower() in ['s', 'si', 'sí', 'y', 'yes']
                    self.consultar_cliente_especifico(cliente, fecha_inicio, fecha_fin, mostrar_detalle)
                elif opcion == '4':
                    fecha_inicio = input("Fecha inicio (YYYY-MM-DD): ").strip()
                    fecha_fin = input("Fecha fin (YYYY-MM-DD): ").strip()
                    self.analizar_periodo(fecha_inicio, fecha_fin)
                elif opcion == '5':
                    titulo = input("Título o programa a buscar: ").strip()
                    fecha_inicio = input("Fecha inicio (YYYY-MM-DD) [Enter para omitir]: ").strip() or None
                    fecha_fin = input("Fecha fin (YYYY-MM-DD) [Enter para omitir]: ").strip() or None
                    self.buscar_por_titulo(titulo, fecha_inicio, fecha_fin)
                elif opcion == '6':
                    self.preguntar_generar_reporte()
                elif opcion == '7':
                    print("🚪 Saliendo del consultor. ¡Hasta luego!")
                    break
                else:
                    print("❌ Opción inválida. Por favor, selecciona una opción del 1 al 7.")
            
            except Exception as e:
                print(f"❌ Error: {e}")
    
    # Métodos para integración con Streamlit
    
    def get_total_records(self):
        """Obtener total de registros para Streamlit"""
        try:
            stats = self.db.obtener_estadisticas_generales()
            return stats['total_emisiones']
        except:
            return 0
    
    def get_unique_clients_count(self):
        """Obtener cantidad de clientes únicos para Streamlit"""
        try:
            stats = self.db.obtener_estadisticas_generales()
            return stats['total_clientes']
        except:
            return 0
    
    def get_today_records_count(self):
        """Obtener registros de hoy para Streamlit"""
        try:
            today = datetime.now().strftime('%Y-%m-%d')
            emisiones = self.db.obtener_emisiones_por_fecha(today, today)
            return len(emisiones)
        except:
            return 0
    
    def get_last_processing_date(self):
        """Obtener fecha del último procesamiento para Streamlit"""
        try:
            stats = self.db.obtener_estadisticas_generales()
            return stats['fecha_fin']
        except:
            return "N/A"
    
    def get_daily_records_chart(self, days=30):
        """Obtener datos para gráfico de registros diarios"""
        try:
            # Calcular fechas
            end_date = datetime.now().date()
            start_date = end_date - timedelta(days=days-1)
            
            # Obtener emisiones del período
            emisiones = self.db.obtener_emisiones_por_fecha(
                start_date.strftime('%Y-%m-%d'),
                end_date.strftime('%Y-%m-%d')
            )
            
            if emisiones.empty:
                return pd.DataFrame()
            
            # Agrupar por fecha
            daily_counts = emisiones.groupby('dia_emision').size().reset_index()
            daily_counts.columns = ['Fecha', 'Emisiones']
            daily_counts['Fecha'] = pd.to_datetime(daily_counts['Fecha'])
            
            return daily_counts
        except:
            return pd.DataFrame()
    
    def get_top_clients_chart(self, top_n=10):
        """Obtener datos para gráfico de top clientes"""
        try:
            df_resumen = self.db.obtener_resumen_por_cliente()
            if df_resumen.empty:
                return pd.DataFrame()
            
            return df_resumen.head(top_n)[['cliente', 'total_emisiones']]
        except:
            return pd.DataFrame()
    
    def get_recent_activity(self, limit=10):
        """Obtener actividad reciente para Streamlit"""
        try:
            # Obtener emisiones de los últimos 7 días
            end_date = datetime.now().date()
            start_date = end_date - timedelta(days=7)
            
            emisiones = self.db.obtener_emisiones_por_fecha(
                start_date.strftime('%Y-%m-%d'),
                end_date.strftime('%Y-%m-%d')
            )
            
            if emisiones.empty:
                return pd.DataFrame()
            
            # Ordenar por fecha y hora más recientes
            emisiones_sorted = emisiones.sort_values(['dia_emision', 'hora_emision'], ascending=[False, False])
            
            # Seleccionar columnas relevantes
            columnas = ['dia_emision', 'hora_emision', 'cliente', 'titulo']
            if 'status' in emisiones.columns:
                columnas.append('status')
            
            return emisiones_sorted[columnas].head(limit)
        except:
            return pd.DataFrame()
    
    def get_all_clients(self):
        """Obtener lista de todos los clientes para Streamlit"""
        try:
            df_resumen = self.db.obtener_resumen_por_cliente()
            if df_resumen.empty:
                return ['Todos']
            
            clientes = ['Todos'] + sorted(df_resumen['cliente'].unique())
            return clientes
        except:
            return ['Todos']
    
    def query_with_filters(self, filtros):
        """Consultar con filtros para Streamlit"""
        try:
            fecha_inicio = filtros.get('fecha_inicio')
            fecha_fin = filtros.get('fecha_fin')
            cliente = filtros.get('cliente')
            
            # Obtener emisiones con filtros
            if cliente and cliente != "Todos":
                emisiones = self.db.consultar_emisiones_por_cliente(cliente, fecha_inicio, fecha_fin)
            else:
                emisiones = self.db.obtener_emisiones_por_fecha(fecha_inicio, fecha_fin)
            
            return emisiones
        except Exception as e:
            print(f"Error en consulta con filtros: {e}")
            return pd.DataFrame()
    
    def generar_reporte_desde_consulta_streamlit(self, config):
        """Generar reporte personalizado - Método específico para Streamlit"""
        try:
            fecha_inicio = config.get('fecha_inicio')
            fecha_fin = config.get('fecha_fin')
            cliente = config.get('cliente')
            formatos_raw = config.get('formatos', ['TXT'])
            
            # Normalizar formatos para compatibilidad
            formatos_normalizados = []
            for formato in formatos_raw:
                if formato.upper() in ['TXT', 'TEXT']:
                    formatos_normalizados.append('TXT')
                elif formato.upper() in ['EXCEL', 'XLSX', 'XLS']:
                    formatos_normalizados.append('EXCEL')
                elif formato.upper() in ['CSV']:
                    formatos_normalizados.append('CSV')
                elif formato.upper() in ['PDF']:
                    formatos_normalizados.append('PDF')
            
            formatos = formatos_normalizados if formatos_normalizados else ['TXT']
            
            # Obtener emisiones con filtros
            if cliente and cliente != "Todos":
                emisiones = self.db.consultar_emisiones_por_cliente(cliente, fecha_inicio, fecha_fin)
            else:
                emisiones = self.db.obtener_emisiones_por_fecha(fecha_inicio, fecha_fin)
            
            if emisiones.empty:
                return {
                    'success': False,
                    'error': 'No se encontraron datos para los filtros especificados',
                    'total_records': 0,
                    'unique_clients': 0,
                    'files': []
                }
            
            # Configurar filtros para el reporte
            filtros = []
            if fecha_inicio:
                filtros.append(f"Desde: {fecha_inicio}")
            if fecha_fin:
                filtros.append(f"Hasta: {fecha_fin}")
            if cliente and cliente != "Todos":
                filtros.append(f"Cliente: {cliente}")
            
            filtros_str = " | ".join(filtros) if filtros else "Sin filtros"
            
            # Definir directorio de reportes
            directorio_reportes = self.directorio_proyecto / "reportes"
            directorio_reportes.mkdir(exist_ok=True)
            
            # Generar nombre base del archivo
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            
            if cliente and cliente != "Todos":
                nombre_base = f"reporte_{cliente.replace(' ', '_')}_{timestamp}"
            else:
                nombre_base = f"reporte_personalizado_{timestamp}"
            
            archivos_generados = []
            unique_clients = emisiones['cliente'].nunique()
            
            # Generar archivos según formatos solicitados
            if 'TXT' in formatos:
                nombre_txt, ruta_txt = self._obtener_nombre_con_version(directorio_reportes, nombre_base, ".txt")
                if self._generar_archivo_txt(emisiones, ruta_txt, filtros_str):
                    archivos_generados.append({
                        'formato': 'TXT',
                        'nombre': nombre_txt,
                        'filename': nombre_txt,
                        'name': nombre_txt,
                        'path': str(ruta_txt),
                        'ruta': str(ruta_txt),
                        'tamaño': ruta_txt.stat().st_size if ruta_txt.exists() else 0
                    })
            
            if 'EXCEL' in formatos:
                nombre_excel, ruta_excel = self._obtener_nombre_con_version(directorio_reportes, nombre_base, ".xlsx")
                if self._generar_archivo_excel(emisiones, ruta_excel):
                    archivos_generados.append({
                        'formato': 'EXCEL',
                        'nombre': nombre_excel,
                        'filename': nombre_excel,
                        'name': nombre_excel,
                        'path': str(ruta_excel),
                        'ruta': str(ruta_excel),
                        'tamaño': ruta_excel.stat().st_size if ruta_excel.exists() else 0
                    })
            
            if 'CSV' in formatos:
                nombre_csv, ruta_csv = self._obtener_nombre_con_version(directorio_reportes, nombre_base, ".csv")
                try:
                    emisiones.to_csv(ruta_csv, index=False, encoding='utf-8')
                    archivos_generados.append({
                        'formato': 'CSV',
                        'nombre': nombre_csv,
                        'filename': nombre_csv,
                        'name': nombre_csv,
                        'path': str(ruta_csv),
                        'ruta': str(ruta_csv),
                        'tamaño': ruta_csv.stat().st_size if ruta_csv.exists() else 0
                    })
                except Exception as e:
                    print(f"Error al generar CSV: {e}")
            
            return {
                'success': True,
                'total_records': len(emisiones),
                'unique_clients': unique_clients,
                'files': archivos_generados,
                'report_config': config
            }
            
        except Exception as e:
            return {
                'success': False,
                'error': f'Error al generar reporte: {str(e)}',
                'total_records': 0,
                'unique_clients': 0,
                'files': []
            }


if __name__ == "__main__":
    try:
        consultor = AsRunConsultor()
        consultor.menu_principal()
    except KeyboardInterrupt:
        print("\n\n👋 Proceso interrumpido por el usuario")
    except Exception as e:
        print(f"\n❌ Error: {e}")
